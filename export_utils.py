import pandas as pd
import io
import hashlib
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def get_color_for_name(name, all_people=None, custom_colors=None):
    """
    Returns a consistent hex color (without #) for a given name.
    If custom_colors is provided and name is in it, use it.
    Otherwise uses a predefined pastel palette based on index.
    """
    if custom_colors and name in custom_colors:
        return custom_colors[name].lstrip("#")
    
    if name == "-" or not isinstance(name, str):
        return "FFFFFF"
        
    # Remove the forced marker so colors stay consistent
    if name.endswith(" (!)"):
        clean_name = name[:-4]
    else:
        clean_name = name
    
    # A curated palette of highly distinct, pleasant pastel/vivid colors
    palette = [
        "FF9999", # Red
        "FFCC99", # Orange
        "FFFF99", # Yellow
        "99FF99", # Green
        "99FFFF", # Cyan
        "9999FF", # Blue
        "CC99FF", # Purple
        "FF99FF", # Magenta
        "FFB366", # Dark Orange
        "00FFCC", # Teal
        "FF99CC", # Pink
        "C2C2A3", # Olive
        "FF6666", # Dark Red
        "66FF66", # Bright Green
        "6666FF"  # Deep Blue
    ]
    
    if all_people and clean_name in all_people:
        color_index = all_people.index(clean_name) % len(palette)
    else:
        # Fallback if list not provided
        import hashlib
        hash_object = hashlib.md5(clean_name.encode())
        hex_hash = hash_object.hexdigest()
        color_index = int(hex_hash, 16) % len(palette)
    
    return palette[color_index]

def get_excel_download(df: pd.DataFrame, stats_df: pd.DataFrame = None, custom_colors: dict = None) -> bytes:
    """
    Generates styled Excel file bytes from a DataFrame with name-based colors.
    """
    # Extract unique people from df for 100% unique colors
    shift_cols = [c for c in ["Tura 1", "Tura 2", "Tura 3", "PZU"] if c in df.columns]
    unique_names = set()
    for col in shift_cols:
        for val in df[col].dropna().unique():
            if isinstance(val, str):
                names = [n.strip() for n in val.replace(" (!)", "").split(",")]
                for n in names:
                    unique_names.add(n)
    unique_names.discard("-")
    all_people = sorted(list(unique_names))
    
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Orar', index=False)
        workbook = writer.book
        worksheet = writer.sheets['Orar']
        
        # Define border style
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
                             
        # Style headers
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
        # Adjust column widths
        for col in range(1, len(df.columns) + 1):
            col_letter = get_column_letter(col)
            worksheet.column_dimensions[col_letter].width = 16
            
        # Style cells
        # Data starts from row 2
        for row in range(2, len(df) + 2):
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                
                # Apply color based on person name if it's a shift column
                col_name = df.columns[col-1]
                if col_name in ["Tura 1", "Tura 2", "Tura 3", "PZU"]:
                    val = str(cell.value) if cell.value else "-"
                    if val != "-":
                        names = [n.strip() for n in val.replace(" (!)", "").split(",")]
                        color_hex = get_color_for_name(names[0], all_people, custom_colors)
                        cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                        
        # Insert empty rows between weeks
        if "Saptamana" in df.columns:
            week_col_idx = df.columns.get_loc("Saptamana") + 1
            current_week = None
            rows_to_insert = []
            for r in range(2, len(df) + 2):
                week_val = worksheet.cell(row=r, column=week_col_idx).value
                if current_week is not None and week_val != current_week:
                    rows_to_insert.append(r)
                current_week = week_val
                
            for r in reversed(rows_to_insert):
                worksheet.insert_rows(r)
                
        # Add Statistics sheet if provided
        if stats_df is not None:
            stats_df.to_excel(writer, sheet_name='Statistici', index=False)
            ws_stats = writer.sheets['Statistici']
            
            for col in range(1, len(stats_df.columns) + 1):
                cell = ws_stats.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws_stats.column_dimensions[get_column_letter(col)].width = 16
                
            for row in range(2, len(stats_df) + 2):
                for col in range(1, len(stats_df.columns) + 1):
                    cell = ws_stats.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

    return output.getvalue()

def get_csv_download(df: pd.DataFrame) -> bytes:
    """
    Generates CSV bytes from a DataFrame.
    """
    return df.to_csv(index=False, encoding='utf-8-sig').encode('utf-8')

def get_multi_month_excel_download(df_dict: dict, combined_stats_df: pd.DataFrame, all_people: list, custom_colors: dict = None) -> bytes:
    """
    Generates an Excel file with multiple sheets (one for each month) and a combined statistics sheet.
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for sheet_name, df in df_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws_col_letter = get_column_letter(col)
                worksheet.column_dimensions[ws_col_letter].width = 16
                
            for row in range(2, len(df) + 2):
                for col in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                    
                    col_name = df.columns[col-1]
                    if col_name in ["Tura 1", "Tura 2", "Tura 3", "PZU"]:
                        val = str(cell.value) if cell.value else "-"
                        if val != "-":
                            names = [n.strip() for n in val.replace(" (!)", "").split(",")]
                            color_hex = get_color_for_name(names[0], all_people, custom_colors)
                            cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                            
            if "Saptamana" in df.columns:
                week_col_idx = df.columns.get_loc("Saptamana") + 1
                current_week = None
                rows_to_insert = []
                for r in range(2, len(df) + 2):
                    week_val = worksheet.cell(row=r, column=week_col_idx).value
                    if current_week is not None and week_val != current_week:
                        rows_to_insert.append(r)
                    current_week = week_val
                    
                for r in reversed(rows_to_insert):
                    worksheet.insert_rows(r)
                    
        # Add Statistics sheet
        if combined_stats_df is not None:
            combined_stats_df.to_excel(writer, sheet_name='Statistici Cumulate', index=False)
            ws_stats = writer.sheets['Statistici Cumulate']
            
            for col in range(1, len(combined_stats_df.columns) + 1):
                cell = ws_stats.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
                ws_stats.column_dimensions[get_column_letter(col)].width = 16
                
            for row in range(2, len(combined_stats_df) + 2):
                for col in range(1, len(combined_stats_df.columns) + 1):
                    cell = ws_stats.cell(row=row, column=col)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border

    return output.getvalue()
